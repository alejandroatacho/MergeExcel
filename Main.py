from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import xlrd

wb = load_workbook('data/MainBook.xlsx')
wb2 = load_workbook('data/NewProducts2.xlsx')

ws = wb.active
ws2 = wb2.active

for row in range(1, 11):
    print(row)
   # for col in range(1, 5):
    #  print(col)
    """char = get_column_letter(col)
        print(ws[char+str(row)])
"""

# wb.save('output/main_test.xlsx')
